"""
Extrae las opciones oficiales del Diccionario UARIV para las preguntas
LISTA / LISTA_MULTIPLE que están con `opciones: []` en las fixtures.

Lee los Excel de docs/perfiles/, busca cada código_externo en las hojas
HOGAR y PERSONAS, e identifica el patrón:

  (a) OPCIONES_CLASICAS:  val=1,2,3,... + descripción → opciones normales
  (b) MATRIZ_SINO:        filas siguientes con B=otro_codigo + val=1 → en
                          realidad son N sub-preguntas binarias
  (c) TEXTO_LIBRE:        val='TEXTO' o 'DIVIPOLA' → no es LISTA realmente
  (d) VACIA:              no se encuentra contenido o solo Campo abierto

Salida: JSON con la clasificación + opciones extraídas para cada pregunta.
"""
import json
import sys
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parents[2]

DICCIONARIOS = {
    'TERRITORIAL': [
        ROOT / 'docs' / 'perfiles' / 'Perfil Territorial_web y offline' /
        'Diccionario_de_datos__Entrevista de Caracterización_V7_Perfil Territorial.xlsx',
    ],
    'TELEFONICO': [
        # Buscar primero en V8 (más reciente) y caer a V7 si no se encuentra
        ROOT / 'docs' / 'perfiles' / 'Perfil Telefonico_SAAH__web' /
        'Diccionario_de_datos__Entrevista de Caracterización_V8_Perfil Telefónico.xlsx',
        ROOT / 'docs' / 'perfiles' / 'Perfil Telefonico_SAAH__web' /
        'Diccionario_de_datos__Entrevista de Caracterización_V7_Perfil Telefonico.xlsx',
        # Fallback final: TERRITORIAL — TELEFONICO comparte muchas preguntas con TERRITORIAL
        ROOT / 'docs' / 'perfiles' / 'Perfil Territorial_web y offline' /
        'Diccionario_de_datos__Entrevista de Caracterización_V7_Perfil Territorial.xlsx',
    ],
}

# Las 34 preguntas LISTA sin opciones identificadas por el QA
PREGUNTAS_VACIAS = {
    'TERRITORIAL': [
        ('C', 'C6A',  'LISTA_MULTIPLE', 'C20'),
        ('C', 'C17A', 'LISTA_MULTIPLE', 'C21'),
        ('D', 'E2',   'LISTA_MULTIPLE', 'D5'),
        ('D', 'RR2',  'LISTA',          'D7'),
        ('D', 'RR10A','LISTA',          'D15'),
        ('E', 'F4',   'LISTA',          'E5'),
        ('B', 'I7A',  'LISTA_MULTIPLE', 'B17'),
        ('B', 'I13',  'LISTA_MULTIPLE', 'B19'),
        ('B', 'B13A', 'LISTA_MULTIPLE', 'B20'),
        ('B', 'A15',  'LISTA',          'B38'),
        ('B', 'A16',  'LISTA',          'B39'),
        ('B', 'A16A', 'LISTA',          'B40'),
        ('G', 'H9',   'LISTA',          'G2'),
        ('G', 'H12A', 'LISTA',          'G5'),
        ('G', 'H14',  'LISTA_MULTIPLE', 'G7'),
        ('H', 'I11A', 'LISTA',          'H3'),
        ('H', 'I25B', 'LISTA',          'H4'),
        ('H', 'I26',  'LISTA',          'H5'),
        ('H', 'I28A', 'LISTA_MULTIPLE', 'H7'),
        ('JF','L6',   'LISTA',          'J6'),
        ('JF','L8',   'LISTA',          'J8'),
        ('K', 'PL1',  'LISTA',          'K2'),
        ('K', 'PL3',  'LISTA',          'K4'),
        ('L', 'FP2',  'LISTA',          'L2'),
        ('L', 'FP3',  'LISTA',          'L3'),
        ('L', 'FP8',  'LISTA_MULTIPLE', 'L8'),
        ('M', 'AT1A', 'LISTA_MULTIPLE', 'M1'),
        ('M', 'AT3',  'LISTA',          'M3'),
        ('M', 'AT5',  'LISTA_MULTIPLE', 'M5'),
    ],
    'TELEFONICO': [
        ('B', 'I7A_tel',  'LISTA_MULTIPLE', 'B12'),
        ('D', 'H14_tel',  'LISTA_MULTIPLE', 'D4'),
        ('F', 'I11A_tel', 'LISTA',          'F3'),
        ('F', 'I25B_tel', 'LISTA',          'F4'),
        ('F', 'I28A_tel', 'LISTA_MULTIPLE', 'F6'),
    ],
}


def analizar_pregunta(ws, codigo_externo: str, codigo_buscar: str = None) -> dict:
    """
    Busca `codigo_externo` en la columna B y analiza las filas siguientes.
    Retorna dict con: encontrada, tipo_detectado, opciones, observacion
    """
    if codigo_buscar is None:
        codigo_buscar = codigo_externo
    # Para TELEFONICO los códigos vienen con sufijo _tel pero el Excel tiene
    # el código base (sin _tel) — probamos ambos.
    candidatos = [codigo_buscar]
    if codigo_buscar.endswith('_tel'):
        candidatos.append(codigo_buscar[:-4])

    fila_pregunta = None
    codigo_en_excel = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        b = row[1] if len(row) > 1 else None
        d = row[3] if len(row) > 3 else None
        # B contiene el codigo, D contiene el texto de la pregunta (no nulo cuando es la fila inicial)
        if b in candidatos and d:
            fila_pregunta = i
            codigo_en_excel = b
            break

    if fila_pregunta is None:
        return {'encontrada': False, 'observacion': 'No encontrada en esta hoja'}

    # Analizar filas siguientes (hasta 30) buscando opciones
    opciones = []
    opciones_texto = []
    tipo_detectado = None
    siguiente_pregunta = None
    matriz_codigos = [codigo_en_excel]

    for j in range(fila_pregunta, min(fila_pregunta + 50, ws.max_row + 1)):
        r = list(ws.iter_rows(min_row=j, max_row=j, values_only=True))[0]
        b = r[1] if len(r) > 1 else None
        d = r[3] if len(r) > 3 else None
        desc = r[5] if len(r) > 5 else None    # F = Descripción
        val = r[6] if len(r) > 6 else None     # G = Valores
        idresp = r[7] if len(r) > 7 else None  # H = ID_RESP

        if j == fila_pregunta:
            # Es la fila de inicio — si tiene val numérico es una opción tambien
            if val and str(val).strip().upper() in ('TEXTO', 'TEXTO_LARGO', 'NUMERICO', 'NUMERIC', 'BOOLEAN', 'FECHA', 'DIVIPOLA'):
                tipo_detectado = 'TEXTO_LIBRE'
                return {
                    'encontrada': True, 'codigo_en_excel': codigo_en_excel,
                    'fila_pregunta': fila_pregunta,
                    'tipo_detectado': tipo_detectado,
                    'observacion': f'Pregunta tipo {val} en Excel — no es LISTA real',
                    'opciones': [],
                }
            # Si val es número, también es opción 1
            if val is not None and desc:
                try:
                    valor_str = str(int(val))
                    opciones.append({'valor': valor_str, 'etiqueta': str(desc).strip(), 'id_resp_vivanto': int(idresp) if idresp else 0, 'orden': len(opciones) + 1})
                    opciones_texto.append(f'{valor_str}: {desc}')
                except (ValueError, TypeError):
                    pass
            continue

        # Si la fila tiene otro codigo distinto en B con texto pregunta D, es la siguiente pregunta
        if b and d and b not in matriz_codigos:
            siguiente_pregunta = (j, b)
            break

        # Si la fila tiene B con codigo distinto SIN texto D, puede ser parte de una matriz Sí/No
        if b and not d and b != codigo_en_excel:
            matriz_codigos.append(b)
            if desc and val == 1:
                opciones_texto.append(f'[MATRIZ: {b}] {desc}')
            continue

        # Fila con B vacío, desc y val — es una opción clásica
        if not b and desc and val is not None:
            try:
                valor_str = str(int(val))
                opciones.append({
                    'valor': valor_str,
                    'etiqueta': str(desc).strip(),
                    'id_resp_vivanto': int(idresp) if idresp else 0,
                    'orden': len(opciones) + 1,
                })
                opciones_texto.append(f'{valor_str}: {desc}')
            except (ValueError, TypeError):
                # val es string como "TEXTO" o "DIVIPOLA"
                if str(val).strip().upper() in ('TEXTO', 'CAMPO ABIERTO', 'DIVIPOLA'):
                    opciones_texto.append(f'[CAMPO LIBRE] {desc}')

    # Para MATRIZ_SINO: re-extraer cada sub-código + su descripción como opción
    if len(matriz_codigos) > 1:
        opciones_matriz = []
        for j in range(fila_pregunta, min(fila_pregunta + 50, ws.max_row + 1)):
            r = list(ws.iter_rows(min_row=j, max_row=j, values_only=True))[0]
            b = r[1] if len(r) > 1 else None
            d = r[3] if len(r) > 3 else None
            desc = r[5] if len(r) > 5 else None
            val = r[6] if len(r) > 6 else None
            idresp = r[7] if len(r) > 7 else None
            # Filas que pertenecen a la matriz: B en matriz_codigos, val=1, desc no vacío
            if b in matriz_codigos and desc and val == 1 and not (j > fila_pregunta and d):
                opciones_matriz.append({
                    'valor': str(b),
                    'etiqueta': str(desc).strip(),
                    'id_resp_vivanto': int(idresp) if idresp else 0,
                    'orden': len(opciones_matriz) + 1,
                })
            # Si llegamos a otra pregunta principal, detener
            if b and d and b not in matriz_codigos:
                break
        if opciones_matriz:
            opciones = opciones_matriz
            tipo_detectado = 'MATRIZ_SINO'
        else:
            tipo_detectado = 'MATRIZ_SINO_VACIA'
    elif opciones:
        tipo_detectado = 'OPCIONES_CLASICAS'
    elif opciones_texto:
        tipo_detectado = 'CAMPO_LIBRE_O_MIXTO'
    else:
        tipo_detectado = 'VACIA'

    return {
        'encontrada': True,
        'codigo_en_excel': codigo_en_excel,
        'fila_pregunta': fila_pregunta,
        'tipo_detectado': tipo_detectado,
        'opciones': opciones,
        'opciones_texto_debug': opciones_texto[:15],
        'matriz_codigos': matriz_codigos if len(matriz_codigos) > 1 else None,
    }


def main():
    resultados = {}

    for perfil, archivos in DICCIONARIOS.items():
        # Cargar todos los workbooks (uno o varios — para fallback)
        workbooks = []
        for archivo in archivos:
            if not archivo.exists():
                print(f'!! No existe: {archivo}', file=sys.stderr)
                continue
            wb = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
            hojas = [s for s in wb.sheetnames if 'Caracter' in s or s in ('HOGAR', 'PERSONAS')]
            workbooks.append((archivo.name, wb, hojas))

        resultados[perfil] = []

        for cap, cod_ext, tipo_actual, no_preg in PREGUNTAS_VACIAS[perfil]:
            # Intentar en cada workbook + cada hoja, en orden
            analisis = {'encontrada': False, 'observacion': 'No encontrada en ningún diccionario'}
            for arch_name, wb, hojas in workbooks:
                for nombre_hoja in hojas:
                    ws = wb[nombre_hoja]
                    a = analizar_pregunta(ws, cod_ext)
                    if a['encontrada']:
                        a['hoja'] = nombre_hoja
                        a['archivo'] = arch_name
                        analisis = a
                        break
                if analisis['encontrada']:
                    break
            resultados[perfil].append({
                'cap': cap,
                'codigo_externo': cod_ext,
                'no_pregunta': no_preg,
                'tipo_actual': tipo_actual,
                'analisis': analisis,
            })

    # Imprimir resumen
    print('\n' + '=' * 80)
    print('RESUMEN DE EXTRACCIÓN')
    print('=' * 80)
    for perfil, items in resultados.items():
        print(f'\n--- {perfil} ({len(items)} preguntas) ---')
        clasif = {'OPCIONES_CLASICAS': 0, 'MATRIZ_SINO': 0, 'TEXTO_LIBRE': 0,
                  'CAMPO_LIBRE_O_MIXTO': 0, 'VACIA': 0, 'NO_ENCONTRADA': 0}
        for item in items:
            a = item['analisis']
            if not a.get('encontrada'):
                clasif['NO_ENCONTRADA'] += 1
                tag = 'NO_ENCONTRADA'
            else:
                tag = a.get('tipo_detectado', '?')
                clasif[tag] = clasif.get(tag, 0) + 1
            n_opc = len(a.get('opciones', []))
            print(f'  {item["cap"]:<3} {item["no_pregunta"]:<5} {item["codigo_externo"]:<12} -> {tag:<22}  opciones={n_opc}')
        print('  Conteo:', clasif)

    # Guardar JSON con toda la info
    out_path = ROOT / 'srni-backend' / 'scripts' / 'opciones_extraidas.json'
    out_path.parent.mkdir(exist_ok=True)
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(resultados, f, indent=2, ensure_ascii=False)
    print(f'\n\nJSON guardado en: {out_path}')


if __name__ == '__main__':
    main()
